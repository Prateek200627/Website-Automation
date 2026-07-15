import re
#
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from dateutil import parser as dateparser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.fbi.gov"
START_URL = "https://www.fbi.gov/wanted/murders"

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.fbi.gov/",
}


# ---------------------------------------------------------------------------
# Step 1: discover every wanted-category link from the section nav
# ---------------------------------------------------------------------------
def get_category_links(start_url):
    response = requests.get(start_url, headers=headers)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    uls = soup.find_all("ul", class_="section-nav-list")
    print(f"Total ULs found: {len(uls)}")

    categories = []
    if len(uls) >= 2:
        second_ul = uls[1]  # Index 1 = second ul
        print("Links:")
        for li in second_ul.find_all("li"):
            a = li.find("a")
            if a and a.get("href"):
                name = a.get_text(strip=True)
                link = a["href"]
                print(name, "->", link)
                categories.append((name, link))
    else:
        print("Second UL not found.")

    return categories


# ---------------------------------------------------------------------------
# Helper: read "Results: 45 Items" from a listing page (for cross-check)
# ---------------------------------------------------------------------------
def get_reported_count(soup_obj):
    div = soup_obj.find("div", class_=lambda c: c and "col-lg-12" in c and "large-12" in c and "hide-for-small" in c)
    if not div:
        return None
    p = div.find("p", class_="right")
    if not p:
        return None
    text = p.get_text(strip=True)
    match = re.search(r"([\d,]+)\s*Items", text, re.IGNORECASE)
    if match:
        return int(match.group(1).replace(",", ""))
    return None


# ---------------------------------------------------------------------------
# Step 2: for one category, walk every listing page (Load More pagination)
# and collect every person link, cross-checking the reported count.
# ---------------------------------------------------------------------------
def get_person_links_for_category(category_url, category_name):
    links = []
    reported_counts = []
    url = category_url

    while url:
        print(f"\n[{category_name}] Reading: {url}\n")

        response = requests.get(url, headers=headers)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, "html.parser")

        # Same selector as: soup.find("ul", class_=lambda c: c and "wanted-grid-natural" in c)
        persons = soup.select("ul.wanted-grid-natural li.portal-type-person")

        for person in persons:
            a = person.find("a", href=True)
            if a:
                links.append(a["href"])
                print(a["href"])

        reported = get_reported_count(soup)
        reported_counts.append(reported)
        print(f"[{category_name}] Reported count: {reported} | Links found so far: {len(links)}")

        button = soup.find("button", class_="load-more")
        if button and button.get("href"):
            url = urljoin(BASE_URL, button["href"])
        else:
            url = None

    final_reported = next((c for c in reversed(reported_counts) if c is not None), None)
    print(f"\n[{category_name}] Total links collected: {len(links)}")
    print(f"[{category_name}] Final reported count: {final_reported}")
    if final_reported is not None:
        print(f"[{category_name}] " + ("MATCH" if final_reported == len(links) else "MISMATCH"))

    return links


# ---------------------------------------------------------------------------
# Helper: format one or more DOB strings into DD/MM/YYYY
# ---------------------------------------------------------------------------
def split_dob(raw_dob):
    if not raw_dob:
        return "", []

    dates = re.findall(r"[A-Za-z]+\s+\d{1,2},\s+\d{4}", raw_dob)

    formatted = []
    for d in dates:
        dt = dateparser.parse(d)
        formatted.append(dt.strftime("%d/%m/%Y"))

    first_dob = formatted[0] if formatted else ""
    remaining_dobs = formatted[1:]

    return first_dob, remaining_dobs
# ---------------------------------------------------------------------------
# Helper: clean an alias string - strip wrapping quotes / stray backslashes
# ---------------------------------------------------------------------------
def clean_alias(alias):
    alias = re.sub(r'[\\"]', "", alias)
    alias = re.sub(r"\s+", " ", alias)
    return alias.strip(' "“”')
# ---------------------------------------------------------------------------
# Step 3: detail-page extractor
# ---------------------------------------------------------------------------
def extract_person_details(person_url, category_name):
    resp = requests.get(person_url, headers=headers)
    resp.raise_for_status()
    psoup = BeautifulSoup(resp.text, "html.parser")

    wrapper = psoup.find("div", class_=lambda c: c and "col-lg-12" in c and "wanted-person-wrapper" in c)
    if not wrapper:
        return None

    data = {"url": person_url, "source_category": category_name}

    # 1. Full name
    h1 = wrapper.find("h1", class_="documentFirstHeading")
    data["full_name"] = h1.get_text(strip=True) if h1 else ""

    # 2. Main summary / main activity
    summary = wrapper.find("p", class_="summary")
    data["main"] = summary.get_text(strip=True) if summary else ""

    # 3. Aliases -> cleaned list, joined with ";"
    alias_div = wrapper.find("div", class_="wanted-person-aliases")
    if alias_div:
        alias_p = alias_div.find("p")
        aliases_raw = alias_p.get_text(" ", strip=True) if alias_p else ""

        aliases = []

        for alias in aliases_raw.split(","):
            alias = clean_alias(alias)

        # Keep only aliases that are pure ASCII (English)
            if alias and alias.isascii():
                aliases.append(alias)

        data["alias"] = "; ".join(aliases)

    else:
        data["alias"] = ""

    # 4/5/6/11/12: fields from the description table + leftover -> details
    dob_raw = ""
    ncic = ""
    gender = ""
    nationality = ""
    other_details = []

    table = wrapper.find("table", class_=lambda c: c and "wanted-person-description" in c)
    if table:
        tbody = table.find("tbody") or table
        for tr in tbody.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) < 2:
                continue
            label = tds[0].get_text(strip=True)
            value = tds[1].get_text(strip=True)

            label_lower = label.lower()
            if label_lower == "date(s) of birth used":
                dob_raw = value
            elif "ncic" in label_lower:
                ncic = value
            elif label_lower == "sex":
                gender = value
            elif label_lower == "nationality":
                nationality = value
            else:
                other_details.append(f"{label} - {value}")

    first_dob, remaining_dobs = split_dob(dob_raw)
    # Add extra DOBs to details
    if remaining_dobs:
        other_details.append(f"Other DOB(s) - {'; '.join(remaining_dobs)}")

    data["dob"] = first_dob
    data["ncic"] = ncic
    data["identity_type"] = "NCIC" if ncic else ""
    data["gender"] = gender
    data["nationality"] = nationality
    data["details"] = "; ".join(other_details)

    # 7. Remarks
    remarks_div = wrapper.find("div", class_="wanted-person-remarks")
    if remarks_div:
        p = remarks_div.find("p")
        data["remarks"] = p.get_text(strip=True) if p else ""
    else:
        data["remarks"] = ""

    # 8. Caution (remark2) - merge every non-empty <p> (including ones
    #    that come after a blank/&nbsp; paragraph) into a single paragraph.
    caution_div = wrapper.find("div", class_="wanted-person-caution")
    if caution_div:
        parts = []
        for p in caution_div.find_all("p"):
            text = p.get_text(strip=True)
            if text and text != "\xa0":
                parts.append(text)
        data["remark2"] = " ".join(parts)
    else:
        data["remark2"] = ""

    # 9. Submit a Tip (remark1) - only the first <p>
    submit_div = wrapper.find("div", class_="wanted-person-submit")
    if submit_div:
        p = submit_div.find("p")
        data["remark1"] = p.get_text(strip=True) if p else ""
    else:
        data["remark1"] = ""

    # 10. Reward (remark3)
    reward_div = wrapper.find("div", class_="wanted-person-reward")
    if reward_div:
        p = reward_div.find("p")
        data["remark3"] = p.get_text(strip=True) if p else ""
    else:
        data["remark3"] = ""

    return data


# ---------------------------------------------------------------------------
# MAIN: discover categories -> crawl each -> extract every person -> Excel
# ---------------------------------------------------------------------------
categories = get_category_links(START_URL)

if not categories:
    # Fallback: at least crawl the starting category if the nav wasn't found
    categories = [("Murder", START_URL)]

all_people = []

for category_name, category_href in categories:
    category_url = urljoin(BASE_URL, category_href)
    print(f"\n{'=' * 70}\nCATEGORY: {category_name} -> {category_url}\n{'=' * 70}")

    person_links = get_person_links_for_category(category_url, category_name)

    for link in person_links:
        full_url = urljoin(BASE_URL, link)
        print(f"\n[{category_name}] Fetching details: {full_url}")
        try:
            details = extract_person_details(full_url, category_name)
            if details:
                all_people.append(details)
            else:
                print("  -> wanted-person-wrapper not found, skipped")
        except requests.RequestException as e:
            print(f"  -> request failed: {e}")

# ---------------------------------------------------------------------------
# Build the Excel workbook: FBI_Wanted
# ---------------------------------------------------------------------------
COLUMNS = [
    "FULL_NAME", "CATEGORY", "F_NAME", "M_NAME", "L_NAME", "GENDER", "DOB",
    "ADD_CITY", "ADD_COUNTRY", "State", "Nationalities", "ADDRESS",
    "Identity Number", "Identity Type", "REF_DATE", "DETAILS", "WEB_LINK",
    "VIOLATION_ID", "SOURCE", "Alias", "Associates", "Main Activity",
    "Citizenship information", "STATUS", "Rem1", "Rem2", "Rem3", "Remarks",
]

wb = Workbook()
ws = wb.active
ws.title = "FBI_Wanted"

header_font = Font(name="Arial", bold=True)
body_font = Font(name="Arial")

for col_idx, col_name in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, person in enumerate(all_people, start=2):
    row_values = {
        "FULL_NAME": person.get("full_name", ""),
        "CATEGORY": "P",
        "F_NAME": "",
        "M_NAME": "",
        "L_NAME": "",
        "GENDER": person.get("gender", ""),
        "DOB": person.get("dob", ""),
        "ADD_CITY": "",
        "ADD_COUNTRY": person.get("nationality", ""),
        "State": "",
        "Nationalities": person.get("nationality", ""),
        "ADDRESS": "",
        "Identity Number": person.get("ncic", ""),
        "Identity Type": person.get("identity_type", ""),
        "REF_DATE": "",
        "DETAILS": person.get("details", ""),
        "WEB_LINK": person.get("url", ""),
        "VIOLATION_ID": "",
        "SOURCE": "FBI",
        "Alias": person.get("alias", ""),
        "Associates": "",
        "Main Activity": person.get("main", ""),
        "Citizenship information": "",
        "STATUS": "",
        "Rem1": person.get("remark1", ""),
        "Rem2": person.get("remark2", ""),
        "Rem3": person.get("remark3", ""),
        "Remarks": person.get("remarks", ""),
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=row_values[col_name])
        cell.font = body_font
        cell.alignment = Alignment(vertical="top", wrap_text=False)

WIDTHS = {
    "FULL_NAME": 26, "CATEGORY": 10, "F_NAME": 10, "M_NAME": 10, "L_NAME": 10,
    "GENDER": 10, "DOB": 22, "ADD_CITY": 12, "ADD_COUNTRY": 14, "State": 10,
    "Nationalities": 14, "ADDRESS": 14, "Identity Number": 16, "Identity Type": 14,
    "REF_DATE": 12, "DETAILS": 45, "WEB_LINK": 40, "VIOLATION_ID": 14,
    "SOURCE": 10, "Alias": 30, "Associates": 14, "Main Activity": 45,
    "Citizenship information": 14, "STATUS": 10, "Rem1": 35, "Rem2": 45,
    "Rem3": 20, "Remarks": 35,
}
for col_idx, col_name in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = WIDTHS.get(col_name, 15)

ws.freeze_panes = "A2"

output_path = "FBI_Wanted.xlsx"
wb.save(output_path)


print(f"\nSaved {len(all_people)} records across {len(categories)} categories to {output_path}")