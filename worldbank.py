import os
import sys
import time
import shutil
import datetime
import re
import pandas as pd
import openpyxl
import unicodedata # 🔥 
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
#
class WBAutomationFramework:
    def __init__(self, site_name, site_url):
        self.site_name = site_name
        self.site_url = site_url
        self.short_prefix = str(self.site_name)[:8].upper().strip()
        
        # ==========================================
        # 1. THE LOCATION ANCHOR
        # ==========================================
        if getattr(sys, 'frozen', False):
            self.base_dir = os.path.dirname(sys.executable)
            exe_name = os.path.basename(sys.executable).lower()
        else:
            self.base_dir = os.path.dirname(os.path.abspath(__file__))
            exe_name = "app.py"

        self.generate_logs = "_withlogs" in exe_name

        # ==========================================
        # 2. DIRECTORY MAPPING & QUARANTINE
        # ==========================================
        self.output_dir = self.base_dir
        self.log_dir = os.path.join(self.base_dir, "Logs")
        self.quarantine_dir = os.path.join(self.base_dir, "_chrome_temp_dl")

    def init_driver(self):
        print("🎛️ Booting Headless Browser...")
        chrome_options = Options()
        # chrome_options.add_argument("--headless") 
        
        os.makedirs(self.quarantine_dir, exist_ok=True)
        
        prefs = {
            "download.default_directory": self.quarantine_dir, 
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        driver = webdriver.Chrome(options=chrome_options)
        return driver

    def wait_for_download(self):
        print("⏳ Waiting for file to land in Quarantine...")
        timeout = 60
        elapsed = 0
        
        while elapsed < timeout:
            if os.path.exists(self.quarantine_dir):
                files = os.listdir(self.quarantine_dir)
                crdownloads = [f for f in files if f.endswith('.crdownload')]
                new_files = [f for f in files if f.endswith('.xlsx') or f.endswith('.xls') or f.endswith('.csv')]
                
                if not crdownloads and new_files:
                    new_files_paths = [os.path.join(self.quarantine_dir, f) for f in new_files]
                    latest_file = max(new_files_paths, key=os.path.getctime)
                    return latest_file
            
            time.sleep(1)
            elapsed += 1
            
        raise TimeoutError("🚨 Download timed out after 60 seconds.")

    def phase_1_download(self, btn_xpath):
        print(f"\n🌐 [Phase 1] Navigating to: {self.site_url}")
        driver = self.init_driver()
        
        try:
            driver.get(self.site_url)
            
            print("🎛️ Hunting for the download button...")
            download_btn = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, btn_xpath))
            )
            
            print("⏳ Letting data grid load before clicking...")
            time.sleep(5) 
            
            driver.execute_script("arguments[0].click();", download_btn)
            
            temp_file_path = self.wait_for_download()
            
            if temp_file_path:
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
                raw_filename = f"{self.site_name} [{timestamp}].xlsx"
                
                final_raw_path = os.path.join(self.base_dir, raw_filename)
                shutil.move(temp_file_path, final_raw_path)
                print(f"✅ Raw file secured at: {final_raw_path}")
                return final_raw_path
            
        except Exception as e:
            print(f"💥 Browser Error: {str(e)}")
            raise e
        finally:
            driver.quit()
            try:
                if os.path.exists(self.quarantine_dir):
                    shutil.rmtree(self.quarantine_dir)
            except:
                pass

    def phase_2_map_template(self, raw_file_path):
        print("\n🧠 [Phase 2] Booting Validation & Mapping Engine...")
        
        new_entries_log = []
        error_log = []
        skipped_count = 0
        total_scanned = 0
        
        try:
            final_filename = f"{self.short_prefix}_Master_Report.xlsx"
            final_path = os.path.join(self.output_dir, final_filename)
            existing_firms = set()
            
            # ==========================================
            # 1. MEMORY VALIDATOR & DYNAMIC GENERATOR
            # ==========================================
            if os.path.exists(final_path):
                memory_df = pd.read_excel(final_path, header=None, skiprows=1) 
                if not memory_df.empty:
                    existing_firms = set(memory_df[0].dropna().astype(str).str.strip())
                
                wb = openpyxl.load_workbook(final_path)
                ws = wb.active
                current_row = ws.max_row + 1 
            else:
                print("🏗️ Master file missing. Generating completely from scratch...")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Wb_Sanctions"
                
                headers = [
                    "FULL_NAME", "CATEGORY", "F_NAME", "M_NAME", "L_NAME",
                    "GENDER", "DOB", "ADD_CITY", "ADD_COUNTRY", "State",
                    "Nationalities", "ADDRESS", "Identity Number", "Identity Type", "REF_DATE",
                    "DETAILS", "WEB_LINK", "VIOLATION_ID", "SOURCE", "Alias", "Associates", 
                    "MainActivity", "Citizenship information",
                    "STATUS", "Rem1", "Rem2", "Rem3", "Remarks"
                ]
                
                for col_num, header_name in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=header_name)
                    
                current_row = 2 

            # ==========================================
            # 2. READ & INJECT RAW DATA (WITH X-RAY VISION)
            # ==========================================
            raw_df = pd.read_excel(raw_file_path, header=None, skiprows=3)
            total_scanned = len(raw_df)
            
            entity_keywords = [
                'LIMITED', 'PRIVATE', 'COMPANY', 'TECHNOLOGY', 'SYSTEMS', 
                'SOFTWARE', 'SOFTWARES', 'TECHNOLOGIES', 'TRADING', 
                'ENGINEERING', 'CONSTRUCTION', 'LLC', '' 'CORPORATION','LIMITED LIABILITY COMPANY',
                'CONSULTING','GROUP','INTERNATIONAL','INCORPORATED'
            ]

            # 🎯 HELPER: DATE FORMATTER
            def format_clean_date(date_val):
                d_str = str(date_val).strip()
                if d_str.lower() in ['nan', 'nat', 'none', '']:
                    return "N/A"
                try:
                    return pd.to_datetime(d_str).strftime("%d/%m/%Y")
                except:
                    return d_str.split(" ")[0]
            
            # 🔥 NEW HELPER: TEXT SANITIZER (Removes weird accents like Ũ)
            def clean_text(text_val):
                t_str = str(text_val).strip()
                if t_str.lower() in ['nan', 'none', '']:
                    return "N/A"
                # This line strips accents: Ũ -> U, é -> e
                return unicodedata.normalize('NFKD', t_str).encode('ascii', 'ignore').decode('utf-8')
            
            print(f"🔍 X-RAY MODE ON: Scanning {total_scanned} rows...")

            for index, row in raw_df.iterrows():
                try:
                    row_data = list(row) + ["N/A"] * 10
                    
                    # Apply text sanitizer instantly on the raw name
                    raw_firm_name = str(row_data[0])
                    firm_name = clean_text(raw_firm_name)
                    
                    if firm_name == 'N/A' or not firm_name:
                        print(f"⚠️ Row {index+4}: Skipped (Blank Name)")
                        continue
                        
                    if firm_name in existing_firms:
                        print(f"⏭️ Row {index+4}: Skipped (Already exists: {firm_name[:15]}...)")
                        skipped_count += 1
                        continue

                    print(f"✅ Row {index+4}: Processing -> {firm_name[:20]}...")

                    gender = "N/A"
                    category = "P" 
                    name_lower = firm_name.lower()
                    
                    if "m/s" in name_lower:
                        category = "E"
                    elif "mrs." in name_lower:
                        gender = "Female"
                    elif "mr." in name_lower:
                        gender = "Male"

                    # ABBREVIATION EXPANDER 
                    firm_name = re.sub(r'\bltd\.?\b', 'LIMITED', firm_name, flags=re.IGNORECASE)
                    firm_name = re.sub(r'\bpvt\.?\b', 'PRIVATE', firm_name, flags=re.IGNORECASE)
                    firm_name = re.sub(r'\bco\.?\b', 'COMPANY', firm_name, flags=re.IGNORECASE)
                    firm_name = re.sub(r'\bllc\.?\b', 'LIMITED LIABILITY COMPANY', firm_name, flags=re.IGNORECASE)
                    firm_name = re.sub(r'\binc\.?\b', 'INCORPORATED', firm_name, flags=re.IGNORECASE)

                    # CATEGORY SORTER
                    name_check = firm_name.upper()
                    for kw in entity_keywords:
                        if kw in name_check:
                            category = "E"
                            break 

                    # Apply text sanitizer to address, country, and grounds
                    address   = clean_text(row_data[2])
                    country   = clean_text(row_data[3])
                    grounds   = clean_text(row_data[6])
                    
                    from_date = format_clean_date(row_data[4])     
                    to_date   = format_clean_date(row_data[5])     
                    
                    details_combined = f"Ineligibility Period {from_date}-{to_date} ; Grounds - {grounds}"

                    ws.cell(row=current_row, column=1, value=firm_name)          
                    ws.cell(row=current_row, column=2, value=category)           
                    ws.cell(row=current_row, column=6, value=gender)             
                    ws.cell(row=current_row, column=9, value=country)            
                    ws.cell(row=current_row, column=12, value=address)           
                    ws.cell(row=current_row, column=16, value=details_combined)  
                    ws.cell(row=current_row, column=17, value=self.site_url)     
                    ws.cell(row=current_row, column=19, value="World Bank Sanction")      
                    
                    new_entries_log.append(firm_name)
                    current_row += 1 
                    
                except Exception as row_error:
                    print(f"❌ CRASH ON ROW {index+4}: {str(row_error)}")
                    error_msg = f"Row {index+4}: {str(row_error)}"
                    error_log.append(error_msg)
                    continue

            wb.save(final_path)
            
        except Exception as e:
            error_log.append(f"CRITICAL PIPELINE ERROR: {str(e)}")
            print(f"\n❌ CRASH DETECTED: {str(e)}")
            
        finally:
            # ==========================================
            # 3. SELF-DESTRUCT RAW FILE 
            # ==========================================
            try:
                if os.path.exists(raw_file_path):
                    # Shredder turned back on so your folders stay clean!
                    os.remove(raw_file_path)
                    print(f"🗑️ Ghost Mode: Downloaded raw file destroyed ({os.path.basename(raw_file_path)})")
            except Exception as e:
                pass

            # ==========================================
            # 4. CONDITIONAL AUDIT LOG
            # ==========================================
            if self.generate_logs:
                os.makedirs(self.log_dir, exist_ok=True)
                
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                log_filename = f"RunLog_{self.short_prefix}_{timestamp}.txt"
                log_filepath = os.path.join(self.log_dir, log_filename)
                
                source_filename = os.path.basename(raw_file_path)
                
                with open(log_filepath, "w", encoding="utf-8") as f:
                    f.write("="*50 + "\n")
                    f.write(f"🤖 AUTOMATION RUN LOG - {self.short_prefix}\n")
                    f.write(f"🕒 Date/Time: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("="*50 + "\n\n")
                    
                    f.write(f"📂 SOURCE FILE USED      : {source_filename}\n")
                    f.write(f"📊 TOTAL ROWS SCANNED    : {total_scanned}\n")
                    f.write(f"⏭️ TOTAL SKIPPED (DUPS)  : {skipped_count}\n")
                    f.write(f"✅ TOTAL NEW ENTRIES     : {len(new_entries_log)}\n")
                    f.write(f"⚠️ TOTAL ERRORS          : {len(error_log)}\n\n")
                    
                    f.write("--- 📝 NEW ENTRIES ADDED ---\n")
                    if new_entries_log:
                        for i, name in enumerate(new_entries_log, 1):
                            f.write(f"{i}. {name}\n")
                    else:
                        f.write("No new entries found today.\n")
                        
                    f.write("\n--- 🚨 ERRORS ENCOUNTERED ---\n")
                    if error_log:
                        for err in error_log:
                            f.write(f"- {err}\n")
                    else:
                        f.write("None. Clean run.\n")
                    
                    f.write("\n" + "="*50)
                    
                print(f"📄 Audit Log successfully generated in Logs folder: {log_filename}")
            else:
                print("🚫 Stealth Mode Active: Logging bypassed (Rename .exe to include '_withlogs' to enable).")

    def run(self, btn_xpath):
        print("-" * 50)
        print(f"🚀 Booting Automation Engine for: {self.site_name}")
        print("-" * 50)
        try:
            raw_file = self.phase_1_download(btn_xpath)
            self.phase_2_map_template(raw_file)
            print("\n🏁 Automation complete. Shutting down engine.")
        except Exception as e:
            print(f"❌ Automation failed: {e}")

if __name__ == "__main__":
    TARGET_SITE = "Sanctioned Firms WB"
    TARGET_URL = "https://www.worldbank.org/en/projects-operations/procurement/debarred-firms"
    DOWNLOAD_BUTTON_XPATH = "//a[@role='button']" 

    bot = WBAutomationFramework(site_name=TARGET_SITE, site_url=TARGET_URL)
    bot.run(btn_xpath=DOWNLOAD_BUTTON_XPATH)