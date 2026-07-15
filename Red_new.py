from curl_cffi import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time
import os
import json
import html
import re
from datetime import datetime
from urllib.parse import urljoin


# STRICT 28-COLUMN TEMPLATE HEADERS
EXCEL_HEADERS = [
    "FULL_NAME", "CATEGORY", "F_NAME", "M_NAME", "L_NAME", "GENDER", "DOB", 
    "ADD_CITY", "ADD_COUNTRY", "State", "Nationalities", "Address", "Identity Number", 
    "Identity Type", "REF_DATE", "DETAILS", "WEB_LINK", "VIOLATION_ID", "SOURCE", 
    "Alias", "Associates", "MainActivity", "Citizenship information", "Status", 
    "Rem1", "Rem2", "Rem3", "Remarks"
]

def safe_excel_save(workbook, filepath):
    while True:
        try:
            workbook.save(filepath)
            break
        except PermissionError:
            print(f"\n   [WARNING] Excel file is open and locked by Windows!")
            print(f"   [ACTION] Please close the file. Retrying in 5 seconds...")
            time.sleep(5)

def parse_name_strict(raw_name):
    parts = str(raw_name).strip().split()
    if not parts: return "", "", ""
    if len(parts) == 1: return parts[0], "", ""
    if len(parts) == 2: return parts[0], "", parts[1]
    return parts[0], " ".join(parts[1:-1]), parts[-1]

def format_dob_and_age(dob_str):
    if not dob_str: return "", "", False
    
    parts = dob_str.replace('-', '/').split('/')
    age_str = ""
    is_just_year = False
    formatted_dob = dob_str
    
    try:
        if len(parts) == 1 and len(parts[0]) == 4:
            year = int(parts[0])
            is_just_year = True
            formatted_dob = parts[0]
        elif len(parts) == 3:
            if len(parts[0]) == 4:
                year = int(parts[0])
                formatted_dob = f"{parts[2]}/{parts[1]}/{parts[0]}"
            elif len(parts[2]) == 4:
                year = int(parts[2])
                formatted_dob = f"{parts[0]}/{parts[1]}/{parts[2]}"
            else:
                year = None
        else:
            year = None
            
        if year:
            age_str = str(datetime.now().year - year)
    except:
        pass
        
    return formatted_dob, age_str, is_just_year

def fetch_reference_data(session):
    print("\n[INFO] Fetching live translation dictionaries from Interpol API...")
    url = "https://www.interpol.int/en/How-we-work/Notices/Red-Notices/View-Red-Notices"
    try:
        response = session.get(url, timeout=30)
        match = re.search(r"data-references='(\{.*?\})'", response.text)
        if match:
            print("   [SUCCESS] Translation dictionaries loaded!")
            return json.loads(html.unescape(match.group(1)))
    except Exception as e:
        print(f"   [ERROR] Error fetching dictionaries: {e}")
    return {}

def run_red_production_scraper():
    cbi_url = "https://cbi.gov.in/interpol-red-notice"
    interpol_base_url = "https://ws-public.interpol.int/notices/v1/red"
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filename = os.path.join(script_dir, "Interpol_Red_Notice_Automated2.xlsx")
    
    session = requests.Session(impersonate="chrome120")
    ref_data = fetch_reference_data(session)
    countries_dict = ref_data.get("countries", {})
    languages_dict = ref_data.get("language", {})
    hair_dict = ref_data.get("hairs", {})        
    eye_dict = ref_data.get("eyesColors", {})    

    existing_links = set()
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        link_col_idx = EXCEL_HEADERS.index("WEB_LINK") + 1
        for row in ws.iter_rows(min_row=2, min_col=link_col_idx, max_col=link_col_idx, values_only=True):
            if row[0]: existing_links.add(str(row[0]).strip())
        print(f"[INFO] Loaded {len(existing_links)} existing links from Excel to skip.")
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADERS)
        safe_excel_save(wb, filename)
        print(f"[INFO] Created a new structured Excel Workbook at: {filename}")

    print("\n" + "="*60)
    print("[START] Fetching Target List from Indian CBI Red Notice Portal...")
    print("="*60 + "\n")
    
    cbi_resp = session.get(cbi_url, verify=False, timeout=30)
    if cbi_resp.status_code != 200:
        print(f"[ERROR] Failed to reach CBI. Status: {cbi_resp.status_code}")
        return
        
    soup = BeautifulSoup(cbi_resp.text, 'html.parser')
    notice_table = soup.find('table', id='commonDataTable') or soup.find('table', class_='commonTable') or soup.find('table')
    if not notice_table:
        print("[ERROR] Could not find ANY table in the CBI HTML.")
        return
        
    rows = notice_table.find('tbody').find_all('tr') if notice_table.find('tbody') else notice_table.find_all('tr')
    print(f"   [INFO] Found {len(rows)} targets. Beginning Extraction...\n")
    
    for index, row in enumerate(rows, 1):
        cols = row.find_all('td')
        if len(cols) < 2: continue
        
        name_tag = cols[1].find('a')
        if name_tag:
            cbi_name = name_tag.text.strip()
            raw_link = name_tag.get('href', '').strip()
        else:
            cbi_name = cols[1].text.strip()
            raw_link = ""
            
        if not cbi_name: continue
        full_link = urljoin("https://cbi.gov.in", raw_link) if raw_link else ""
        if full_link and full_link in existing_links: continue

        row_raw_html = str(row)
        id_match = re.search(r'((?:19|20)\d{2})[-/](\d{3,7})', row_raw_html)
        
        identity_number = ""
        if id_match:
            identity_number = f"{id_match.group(1)}-{id_match.group(2)}"
            
        api_success = False

        if identity_number:
            retry_count = 0
            while retry_count < 3:
                try:
                    detail_url = f"{interpol_base_url}/{identity_number}"
                    api_resp = session.get(detail_url, timeout=20)
                    
                    if api_resp.status_code == 200:
                        d = api_resp.json()
                        
                        raw_nats = d.get('nationalities', [])
                        if not isinstance(raw_nats, list): raw_nats = [raw_nats] if raw_nats else []
                        nationalities_str = ", ".join([countries_dict.get(nat, nat) for nat in raw_nats])
                        
                        raw_langs = d.get('languages_spoken_ids', [])
                        if not isinstance(raw_langs, list): raw_langs = [raw_langs] if raw_langs else []
                        languages_str = "/ ".join([languages_dict.get(lang, lang) for lang in raw_langs])

                        hair_str = ", ".join([hair_dict.get(h, h) for h in (d.get('hairs_id') or [])])
                        eyes_str = ", ".join([eye_dict.get(e, e) for e in (d.get('eyes_colors_id') or [])])
                        
                        arrest_warrants = d.get('arrest_warrants', [])
                        charges_list = [w.get('charge') for w in arrest_warrants if w.get('charge')]
                        charges_str = " ; ".join(charges_list)
                        
                        # 🎯 ADD_COUNTRY mirrors Nationality exactly
                        add_country_val = nationalities_str
                        
                        dob_raw = d.get('date_of_birth', '')
                        formatted_dob, age_str, is_just_year = format_dob_and_age(dob_raw)
                        
                        final_dob_col = ""
                        add_dob_to_details = False
                        if dob_raw:
                            if is_just_year:
                                add_dob_to_details = True
                            else:
                                final_dob_col = formatted_dob
                                
                        marks_str = d.get('distinguishing_marks', '')
                        
                        details_parts = []
                        if add_dob_to_details: details_parts.append(f"Date of birth - {formatted_dob}")
                        if age_str: details_parts.append(f"AGE - {age_str}")
                        if marks_str: details_parts.append(f"Distinguishing marks and characteristics-{marks_str}")
                        if d.get('height'): details_parts.append(f"Height-{d.get('height')} metres")
                        if d.get('weight'): details_parts.append(f"Weight-{d.get('weight')} kilograms")
                        if eyes_str: details_parts.append(f"Colour of eyes - {eyes_str}")
                        if hair_str: details_parts.append(f"Colour of hair - {hair_str}")
                        if languages_str: details_parts.append(f"LANGUAGE(S) SPOKEN - {languages_str}")
                        if charges_str: details_parts.append(f"Charges - {charges_str}")
                        
                        details_string = " ; ".join(details_parts)
                        
                        final_full_name = f"{d.get('forename', '')} {d.get('name', '')}".strip() or cbi_name
                        f_name, m_name, l_name = parse_name_strict(final_full_name)
                        
                        aliases_raw = d.get('aliases', [])
                        alias_list = []
                        for a in aliases_raw:
                            if isinstance(a, dict):
                                a_name = f"{a.get('forename', '')} {a.get('name', '')}".strip()
                                if a_name: alias_list.append(a_name)
                            elif isinstance(a, str):
                                alias_list.append(a)
                        alias_str = ", ".join(alias_list)
                        
                        pob = d.get('place_of_birth', '')
                        cob_id = d.get('country_of_birth_id', '')
                        cob = countries_dict.get(cob_id, cob_id)
                        rem1_str = f"Place of birth - {pob} {cob}".strip() if pob or cob else ""
                        
                        record = {
                            "FULL_NAME": final_full_name,
                            "CATEGORY": "P",
                            "F_NAME": f_name,
                            "M_NAME": m_name,
                            "L_NAME": l_name,
                            "GENDER": "FEMALE" if d.get('sex_id') == "F" else "MALE", 
                            "DOB": final_dob_col,
                            "ADD_CITY": "",
                            "ADD_COUNTRY": add_country_val,
                            "State": "",
                            "Nationalities": nationalities_str,
                            "Address": "",
                            "Identity Number": identity_number,
                            "Identity Type": "",
                            "REF_DATE": "",
                            "DETAILS": details_string,
                            "WEB_LINK": full_link,
                            "VIOLATION_ID": "",
                            "SOURCE": "Interpol Red Notices",
                            "Alias": alias_str,
                            "Associates": "",
                            "MainActivity": "",
                            "Citizenship information": "",
                            "Status": "",
                            "Rem1": rem1_str,
                            "Rem2": "", 
                            "Rem3": "",
                            "Remarks": ""
                        }
                        
                        row_data = [record.get(col, "") for col in EXCEL_HEADERS]
                        ws.append(row_data)
                        safe_excel_save(wb, filename)
                        
                        existing_links.add(full_link)
                        api_success = True
                            
                        print(f"   [SUCCESS] [{index}] API Red Data Extracted: {record['FULL_NAME']}")
                        break 
                        
                    elif api_resp.status_code in [403, 429]:
                        time.sleep(5)
                        retry_count += 1
                    else:
                        print(f"   [WARN] [{index}] API Error {api_resp.status_code} for ID: {identity_number}. Falling back.")
                        break
                        
                except Exception as e:
                    time.sleep(5)
                    retry_count += 1
                    
            time.sleep(0.5)
            
        # ==========================================
        # PATH B: THE FALLBACK NAME-ONLY ROUTE
        # ==========================================
        if not api_success:
            f_name, m_name, l_name = parse_name_strict(cbi_name)
            
            record = {
                "FULL_NAME": cbi_name,
                "CATEGORY": "P",
                "F_NAME": f_name,
                "M_NAME": m_name,
                "L_NAME": l_name,
                "WEB_LINK": full_link,
                "SOURCE": "Interpol Red Notices"
            }
            
            row_data = [record.get(col, "") for col in EXCEL_HEADERS]
            ws.append(row_data)
            safe_excel_save(wb, filename)
            
            existing_links.add(full_link)
            print(f"   [SUCCESS] [{index}] Basic Red Name Extracted: {cbi_name}")

    print("\n[COMPLETE] Script Complete! Excel file updated.")

if __name__ == "__main__":
    run_red_production_scraper()