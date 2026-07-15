from curl_cffi import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time
import os
import json
import html
import re
from datetime import datetime
from urllib.parse import urljoin
#

# STRICT 28-COLUMN TEMPLATE HEADERS
EXCEL_HEADERS = [
    "FULL_NAME", "CATEGORY", "F_NAME", "M_NAME", "L_NAME", "GENDER", "DOB", 
    "ADD_CITY", "ADD_COUNTRY", "State", "Nationalities", "ADDRESS", "Identity Number", 
    "Identity Type", "REF_DATE", "DETAILS", "WEB_LINK", "VIOLATION_ID", "SOURCE", 
    "Alias", "Associates", "MainActivity", "Citizenship information", "STATUS", 
    "Rem1", "Rem2", "Rem3", "Remarks"
]

def safe_excel_save(workbook, filepath):
    while True:
        try:
            workbook.save(filepath)
            break
        except PermissionError:
            print(f"\n   [WARNING] Excel file is currently open and locked by Windows!")
            print(f"   [ACTION] Please close the file. Retrying save in 5 seconds...")
            time.sleep(5)

def parse_name_strict(raw_name):
    """Splits name accurately into First, Middle, and Last."""
    parts = str(raw_name).strip().split()
    if not parts:
        return "", "", ""
    if len(parts) == 1:
        return parts[0], "", ""
    if len(parts) == 2:
        return parts[0], "", parts[1]
    
    f_name = parts[0]
    l_name = parts[-1]
    m_name = " ".join(parts[1:-1])
    return f_name, m_name, l_name

def fetch_reference_data(session):
    print("\n[INFO] Fetching live translation dictionaries from Interpol API...")
    url = "https://www.interpol.int/en/How-we-work/Notices/Yellow-Notices/View-Yellow-Notices"
    try:
        response = session.get(url, timeout=30)
        match = re.search(r"data-references='(\{.*?\})'", response.text)
        if match:
            print("   [SUCCESS] Translation dictionaries loaded!")
            return json.loads(html.unescape(match.group(1)))
    except Exception as e:
        print(f"   [ERROR] Error fetching dictionaries: {e}")
    return {}

def run_final_update_scraper():
    cbi_url = "https://cbi.gov.in/interpol-yellow-notice"
    interpol_base_url = "https://ws-public.interpol.int/notices/v2/yellow"
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filename = os.path.join(script_dir, "New_Yellow.xlsx")
    
    session = requests.Session(impersonate="chrome120")
    
    ref_data = fetch_reference_data(session)
    countries_dict = ref_data.get("countries", {})
    languages_dict = ref_data.get("language", {})
    hair_dict = ref_data.get("hairs", {})        
    eye_dict = ref_data.get("eyesColors", {})    

    existing_links = set()
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        link_col_idx = EXCEL_HEADERS.index("WEB_LINK") + 1
        
        for row in ws.iter_rows(min_row=2, min_col=link_col_idx, max_col=link_col_idx, values_only=True):
            if row[0]:
                existing_links.add(str(row[0]).strip())
                
        print(f"[INFO] Loaded {len(existing_links)} existing links from Excel to skip.")
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADERS)
        safe_excel_save(wb, filename)
        print(f"[INFO] Created a new structured Excel Workbook at: {filename}")

    print("\n" + "="*60)
    print("[START] Fetching Target List from Indian CBI Portal...")
    print("="*60 + "\n")
    
    cbi_resp = session.get(cbi_url, verify=False, timeout=30)
    if cbi_resp.status_code != 200:
        print(f"[ERROR] Failed to reach CBI. Status: {cbi_resp.status_code}")
        return
        
    soup = BeautifulSoup(cbi_resp.text, 'html.parser')
    
    notice_table = soup.find('table', class_='notice-table')
    if not notice_table:
        notice_table = soup.find('table')
        
    if not notice_table:
        print("[ERROR] Could not find ANY table in the CBI HTML.")
        return
        
    rows = notice_table.find('tbody').find_all('tr') if notice_table.find('tbody') else notice_table.find_all('tr')
    print(f"   [INFO] Found {len(rows)} targets. Beginning Incremental Extraction...\n")
    
    for index, row in enumerate(rows, 1):
        cols = row.find_all('td')
        if len(cols) < 2: 
            continue
        
        name_tag = cols[1].find('a')
        
        if name_tag:
            cbi_name = name_tag.text.strip()
            raw_link = name_tag.get('href', '')
        else:
            cbi_name = cols[1].text.strip()
            raw_link = ""
            
        if not cbi_name:
            continue
            
        full_link = urljoin("https://cbi.gov.in", raw_link) if raw_link else ""
        
        if full_link and full_link in existing_links:
            continue

        id_match = re.search(r'(\d{4}[-/]\d+)', full_link) if full_link else None
        identity_number = id_match.group(1).replace("/", "-") if id_match else ""

        api_success = False

        # ==========================================
        # PATH A: THE INTERPOL API ROUTE
        # ==========================================
        if identity_number:
            retry_count = 0
            while retry_count < 3:
                try:
                    detail_url = f"{interpol_base_url}/{identity_number}"
                    api_resp = session.get(detail_url, timeout=20)
                    
                    if api_resp.status_code == 200:
                        d = api_resp.json()
                        
                        raw_nats = d.get('nationalities', [])
                        if not isinstance(raw_nats, list): raw_nats = [raw_nats] if raw_nats else []
                        nationalities_str = ", ".join([countries_dict.get(nat, nat) for nat in raw_nats])
                        
                        translated_country = countries_dict.get(d.get('country', ''), d.get('country', ''))
                        
                        raw_langs = d.get('languages_spoken_ids', [])
                        if not isinstance(raw_langs, list): raw_langs = [raw_langs] if raw_langs else []
                        languages_str = ", ".join([languages_dict.get(lang, lang) for lang in raw_langs])

                        hair_str = ", ".join([hair_dict.get(h, h) for h in (d.get('hairs_id') or [])])
                        eyes_str = ", ".join([eye_dict.get(e, e) for e in (d.get('eyes_colors_id') or [])])
                        
                        details_string = f"Place of disappearance - {d.get('place', '')} ; Date of disappearance - {d.get('date_of_event', '')} ; Height - {d.get('height', '')} ; Weight - {d.get('weight', '')} ; Hair - {hair_str} ; Eyes - {eyes_str} ; Languages - {languages_str}".strip(" ; ")
                        
                        final_full_name = f"{d.get('forename', '')} {d.get('name', '')}".strip() or cbi_name
                        f_name, m_name, l_name = parse_name_strict(final_full_name)
                        
                        record = {
                            "FULL_NAME": final_full_name,
                            "CATEGORY": "P",
                            "F_NAME": f_name,
                            "M_NAME": m_name,
                            "L_NAME": l_name,
                            "GENDER": "Female" if d.get('sex_id') == "F" else "Male",
                            "DOB": d.get('date_of_birth', ''),
                            "ADD_CITY": "",
                            "ADD_COUNTRY": translated_country,
                            "State": "",
                            "Nationalities": nationalities_str,
                            "ADDRESS": "",
                            "Identity Number": identity_number,
                            "Identity Type": "",
                            "REF_DATE": "",
                            "DETAILS": details_string,
                            "WEB_LINK": full_link,
                            "VIOLATION_ID": "",
                            "SOURCE": "Interpol",
                            "Alias": "",
                            "Associates": "",
                            "MainActivity": "",
                            "Citizenship information": "",
                            "STATUS": "",
                            "Rem1": f"Place of birth - {d.get('place_of_birth', '')}" if d.get('place_of_birth') else "",
                            "Rem2": f"Father - {d.get('father_forename', '')} {d.get('father_name', '')} ; Mother - {d.get('mother_forename', '')} {d.get('mother_name', '')}".strip(" ; "),
                            "Rem3": "",
                            "Remarks": f"Last updated {datetime.today().strftime('%d/%b/%y')}"
                        }
                        
                        row_data = [record.get(col, "") for col in EXCEL_HEADERS]
                        ws.append(row_data)
                        safe_excel_save(wb, filename)
                        
                        existing_links.add(full_link)
                        api_success = True
                            
                        print(f"   [SUCCESS] [{index}] API Data Extracted: {record['FULL_NAME']}")
                        break 
                        
                    elif api_resp.status_code in [403, 429]:
                        time.sleep(5)
                        retry_count += 1
                    else:
                        print(f"   [WARN] [{index}] API Error {api_resp.status_code} for ID: {identity_number}. Falling back to name layout.")
                        break
                        
                except Exception as e:
                    time.sleep(5)
                    retry_count += 1
                    
            time.sleep(0.5)
            
        # ==========================================
        # PATH B: THE FALLBACK NAME-ONLY ROUTE 
        # ==========================================
        if not api_success:
            f_name, m_name, l_name = parse_name_strict(cbi_name)
            
            record = {
                "FULL_NAME": cbi_name,
                "CATEGORY": "P",
                "F_NAME": f_name,
                "M_NAME": m_name,
                "L_NAME": l_name,
                "WEB_LINK": full_link,
                "SOURCE": "Interpol",
                "Remarks": f"Last updated {datetime.today().strftime('%d/%b/%y')}"
            }
            
            row_data = [record.get(col, "") for col in EXCEL_HEADERS]
            ws.append(row_data)
            safe_excel_save(wb, filename)
            
            existing_links.add(full_link)
            print(f"   [SUCCESS] [{index}] Basic Name Extracted: {cbi_name}")

    print("\n[COMPLETE] Incremental Update Complete! Excel file updated.")

if __name__ == "__main__":
    run_final_update_scraper()